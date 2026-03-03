"""
Build vector database from AHTN Excel file for RAG-based HS code classification.
Pillar 2: Liability Shield.

Usage:
  1. Set OPENAI_API_KEY in .env
  2. Run: python -m scripts.pillar2.build_ahtn_vector_db

Output:
  - backend/data/pillar2/ahtn_embeddings.npy
  - backend/data/pillar2/ahtn_metadata.json
"""

import json
import os
import re
from pathlib import Path

import numpy as np
from dotenv import load_dotenv
from openai import OpenAI
from openpyxl import load_workbook

# Project root (from backend/scripts/pillar2/)
ROOT_DIR = Path(__file__).resolve().parent.parent.parent.parent
load_dotenv(ROOT_DIR / ".env")

EMBEDDING_MODEL = os.getenv("AHTN_EMBEDDING_MODEL", "text-embedding-3-small")
AHTN_EXCEL_PATH = ROOT_DIR / "AHTN" / "ATIGA-Tariff-Schedules-Malaysia-AHTN-2022-endorsed-AFTAC36.xlsx"
DATA_DIR = ROOT_DIR / "backend" / "data" / "pillar2"
EMBEDDINGS_PATH = DATA_DIR / "ahtn_embeddings.npy"
METADATA_PATH = DATA_DIR / "ahtn_metadata.json"


def _clean(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


CHAPTER_KEYWORDS: dict[str, str] = {
    "84": "machinery machine engine computer laptop",
    "85": "electrical electronic phone smartphone mobile headphone earphone cable wire battery wireless bluetooth usb charging data conductor transmission apparatus",
    "39": "plastic polymer container",
    "42": "bag handbag luggage travel leather",
    "61": "knitted apparel shirt t-shirt blouse",
    "62": "woven apparel trousers jeans pants",
    "18": "cocoa chocolate",
    "19": "biscuit bread pastry cereal",
    "20": "preserved fruit vegetable canned jam",
    "09": "coffee tea spice",
    "17": "sugar sweetener",
    "40": "rubber latex glove tyre",
    "69": "ceramic pottery mug tableware",
    "44": "wood bamboo board",
    "82": "knife cutlery tool blade",
    "30": "medicine pharmaceutical tablet medicament",
}


def _get_chapter_keywords(chapter: str) -> str:
    return CHAPTER_KEYWORDS.get(chapter, "")


def _is_ahtn_code(val) -> bool:
    if val is None:
        return False
    s = str(val).strip()
    return bool(re.match(r"^\d{4}\.\d{2}(\.\d{2})?$", s))


def parse_ahtn_excel(excel_path: Path) -> list[dict]:
    wb = load_workbook(excel_path, read_only=True)
    ws = wb.active

    records = []
    current_heading = None
    description_stack: list[str] = []

    for row in ws.iter_rows(values_only=True):
        if not row:
            continue

        col_a = _clean(row[0])
        col_b = _clean(row[1])
        col_c = _clean(row[2])
        col_d = _clean(row[3])
        col_e = row[4]

        if col_a and re.match(r"^\d{2}\.\d{2}$", col_a):
            current_heading = col_a
            description_stack = []

        if col_c:
            stripped = col_c.lstrip("- ")
            dashes = len(col_c) - len(col_c.lstrip("-"))
            level = dashes // 2 if dashes else 0
            description_stack = description_stack[: level] + [stripped]

        if col_b and _is_ahtn_code(col_b):
            full_desc = " > ".join(description_stack) if description_stack else (col_c or "")
            rate = col_e if col_e is not None else 0
            rate_str = f"{rate}%" if isinstance(rate, (int, float)) else str(rate)

            chapter = col_b[:2] if len(col_b) >= 2 else ""
            chapter_keywords = _get_chapter_keywords(chapter)
            enriched_desc = f"{chapter_keywords} {full_desc}" if chapter_keywords else full_desc

            records.append({
                "ahtn_code": col_b,
                "description": full_desc,
                "heading": current_heading or "",
                "unit": col_d or "",
                "rate": rate_str,
                "full_context": f"AHTN {col_b}: {enriched_desc}. Unit: {col_d or 'N/A'}. Import duty: {rate_str}.",
            })

    wb.close()
    return records


def build_vector_db(records: list[dict]) -> None:
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        raise ValueError("OPENAI_API_KEY not set in .env")

    client = OpenAI(api_key=api_key)
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    batch_size = 100
    texts = [r["full_context"] for r in records]
    all_embeddings = []

    for i in range(0, len(texts), batch_size):
        batch_texts = texts[i : i + batch_size]
        response = client.embeddings.create(
            model=EMBEDDING_MODEL,
            input=batch_texts,
        )
        batch_embeddings = [e.embedding for e in sorted(response.data, key=lambda x: x.index)]
        all_embeddings.extend(batch_embeddings)
        print(f"  Indexed {min(i + batch_size, len(texts))} / {len(texts)} records")

    embeddings = np.array(all_embeddings, dtype=np.float32)
    np.save(EMBEDDINGS_PATH, embeddings)

    metadata = [
        {"ahtn_code": r["ahtn_code"], "description": r["description"], "heading": r["heading"], "unit": r["unit"], "rate": r["rate"]}
        for r in records
    ]
    with open(METADATA_PATH, "w", encoding="utf-8") as f:
        json.dump(metadata, f, ensure_ascii=False, indent=0)

    print(f"\nVector DB built:")
    print(f"  Embeddings: {EMBEDDINGS_PATH}")
    print(f"  Metadata:   {METADATA_PATH}")
    print(f"  Total records: {len(records)}")


def main() -> None:
    import sys
    parse_only = "--parse-only" in sys.argv

    if not AHTN_EXCEL_PATH.exists():
        print(f"Error: AHTN Excel not found at {AHTN_EXCEL_PATH}")
        return

    print("Parsing AHTN Excel...")
    records = parse_ahtn_excel(AHTN_EXCEL_PATH)
    print(f"  Parsed {len(records)} AHTN tariff lines")

    if parse_only:
        print("\n(Parse-only mode: skipping embeddings.)")
        for r in records[:3]:
            print(f"  {r['ahtn_code']}: {r['description'][:60]}...")
        return

    print("\nBuilding vector database (embedding with OpenAI)...")
    build_vector_db(records)
    print("Done.")


if __name__ == "__main__":
    main()
